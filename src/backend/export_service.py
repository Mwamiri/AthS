"""
Data Export Service
Generate CSV and Excel exports for athletes, races, results, and reports
"""

import csv
import io
import logging
from typing import List, Dict, Any, Optional, Tuple
from datetime import datetime
import json
from flask import send_file

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False


class ExportService:
    """Handle data export in various formats"""
    
    def __init__(self):
        self.logger = logging.getLogger('athsys.export')
    
    def export_to_csv(self, data: List[Dict[str, Any]], columns: Optional[List[str]] = None) -> Tuple[io.StringIO, str]:
        """
        Export data to CSV format
        
        Args:
            data: List of dictionaries
            columns: Specific columns to export (None = all)
            
        Returns:
            Tuple of (CSV buffer, filename)
        """
        if not data:
            return None, "No data to export"
        
        try:
            # Determine columns
            if columns is None:
                columns = list(data[0].keys())
            
            # Create CSV
            output = io.StringIO()
            writer = csv.DictWriter(output, fieldnames=columns)
            
            writer.writeheader()
            for row in data:
                writer.writerow({col: row.get(col, '') for col in columns})
            
            filename = f"export_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.csv"
            
            self.logger.info(f"CSV export created: {filename}")
            return output, filename
            
        except Exception as e:
            self.logger.error(f"CSV export failed: {str(e)}")
            raise
    
    def export_to_excel(self, data: List[Dict[str, Any]], sheet_name: str = "Data", 
                       columns: Optional[List[str]] = None, title: Optional[str] = None) -> Tuple[io.BytesIO, str]:
        """
        Export data to Excel format
        
        Args:
            data: List of dictionaries
            sheet_name: Excel sheet name
            columns: Specific columns to export
            title: Optional title row
            
        Returns:
            Tuple of (Excel buffer, filename)
        """
        if not EXCEL_AVAILABLE:
            raise ImportError("openpyxl not installed. Install with: pip install openpyxl")
        
        if not data:
            return None, "No data to export"
        
        try:
            from openpyxl import Workbook
            from openpyxl.utils import get_column_letter
            
            wb = Workbook()
            ws = wb.active
            ws.title = sheet_name
            
            # Add title if provided
            if title:
                ws.append([title])
                ws[f'A1'].font = Font(bold=True, size=14)
                ws.append([])
            
            # Determine columns
            if columns is None:
                columns = list(data[0].keys())
            
            # Write headers
            header_row = ws.append(columns)
            for cell in header_row:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="1F4788", end_color="1F4788", fill_type="solid")
                cell.alignment = Alignment(horizontal="center")
            
            # Write data rows
            for row_data in data:
                row = [row_data.get(col, '') for col in columns]
                ws.append(row)
            
            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Write to buffer
            output = io.BytesIO()
            wb.save(output)
            output.seek(0)
            
            filename = f"export_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.xlsx"
            
            self.logger.info(f"Excel export created: {filename}")
            return output, filename
            
        except Exception as e:
            self.logger.error(f"Excel export failed: {str(e)}")
            raise
    
    def export_athletes_csv(self, athletes: List[Dict]) -> Tuple[io.StringIO, str]:
        """Export athletes to CSV"""
        columns = ['id', 'user_id', 'first_name', 'last_name', 'date_of_birth', 
                   'gender', 'country', 'club', 'category', 'created_at', 'updated_at']
        return self.export_to_csv(athletes, columns)
    
    def export_athletes_excel(self, athletes: List[Dict]) -> Tuple[io.BytesIO, str]:
        """Export athletes to Excel"""
        columns = ['id', 'user_id', 'first_name', 'last_name', 'date_of_birth', 
                   'gender', 'country', 'club', 'category', 'created_at', 'updated_at']
        return self.export_to_excel(athletes, "Athletes", columns, "Athletes Report")
    
    def export_races_csv(self, races: List[Dict]) -> Tuple[io.StringIO, str]:
        """Export races to CSV"""
        columns = ['id', 'name', 'description', 'date', 'location', 'distance', 
                   'category', 'max_participants', 'status', 'created_at']
        return self.export_to_csv(races, columns)
    
    def export_races_excel(self, races: List[Dict]) -> Tuple[io.BytesIO, str]:
        """Export races to Excel"""
        columns = ['id', 'name', 'description', 'date', 'location', 'distance', 
                   'category', 'max_participants', 'status', 'created_at']
        return self.export_to_excel(races, "Races", columns, "Races Report")
    
    def export_results_csv(self, results: List[Dict]) -> Tuple[io.StringIO, str]:
        """Export race results to CSV"""
        columns = ['id', 'race_id', 'athlete_id', 'position', 'time_seconds', 
                   'pace', 'status', 'created_at']
        return self.export_to_csv(results, columns)
    
    def export_results_excel(self, results: List[Dict]) -> Tuple[io.BytesIO, str]:
        """Export race results to Excel"""
        columns = ['id', 'race_id', 'athlete_id', 'position', 'time_seconds', 
                   'pace', 'status', 'created_at']
        return self.export_to_excel(results, "Results", columns, "Race Results Report")
    
    def export_registrations_csv(self, registrations: List[Dict]) -> Tuple[io.StringIO, str]:
        """Export race registrations to CSV"""
        columns = ['id', 'race_id', 'athlete_id', 'bib_number', 'category', 
                   'status', 'registered_at', 'checked_in_at']
        return self.export_to_csv(registrations, columns)
    
    def export_registrations_excel(self, registrations: List[Dict]) -> Tuple[io.BytesIO, str]:
        """Export race registrations to Excel"""
        columns = ['id', 'race_id', 'athlete_id', 'bib_number', 'category', 
                   'status', 'registered_at', 'checked_in_at']
        return self.export_to_excel(registrations, "Registrations", columns, "Race Registrations Report")
    
    def generate_report_csv(self, report_data: Dict[str, List[Dict]], filename: Optional[str] = None) -> Tuple[io.StringIO, str]:
        """Generate multi-section report as CSV"""
        try:
            output = io.StringIO()
            
            for section_name, section_data in report_data.items():
                output.write(f"\n{section_name}\n")
                if section_data:
                    columns = list(section_data[0].keys())
                    writer = csv.DictWriter(output, fieldnames=columns)
                    writer.writeheader()
                    for row in section_data:
                        writer.writerow({col: row.get(col, '') for col in columns})
                output.write("\n")
            
            fname = filename or f"report_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.csv"
            self.logger.info(f"Report CSV created: {fname}")
            return output, fname
            
        except Exception as e:
            self.logger.error(f"Report CSV generation failed: {str(e)}")
            raise


def get_export_service() -> ExportService:
    """Get export service instance"""
    return ExportService()
